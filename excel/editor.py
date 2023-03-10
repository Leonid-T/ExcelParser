from abc import ABC, abstractmethod
from openpyxl import Workbook
from openpyxl.styles import Alignment
from openpyxl.styles.borders import Side, Border


class ExcelEditor(ABC):
    def __init__(self):
        self._wb = Workbook()
        self._sheet = self._wb.active

    @abstractmethod
    def data_to_excel(self, data):
        pass

    def save(self, filepath):
        self._wb.save(filepath)


class ExcelTotalEditor(ExcelEditor):
    def data_to_excel(self, data):
        """
        Packing data into excel file
        """
        self._create_title()
        for row in data:
            self._sheet.append(row)

    def _create_title(self):
        """
        Creating a titles in the header of an excel file
        """
        titles = ["Total", "fact", "forecast"]
        titles.extend(2 * ["Qliq", "Qoil"])
        titles.extend(4 * ["data1", "data2"])

        i = 0
        bs = Side("thin")
        border = Border(left=bs, right=bs, top=bs, bottom=bs)
        center_alignment = Alignment(horizontal='center')
        for row in range(1, 5):
            width = 2 ** (4 - row) - 1
            for coll in range(1, 9, width + 1):
                self._sheet.cell(row=row, column=coll).value = titles[i]
                self._sheet.cell(row=row, column=coll).alignment = center_alignment
                self._sheet.cell(row=row, column=coll).border = border
                self._sheet.merge_cells(start_row=row, end_row=row, start_column=coll, end_column=coll + width)
                i += 1

        self._sheet["I1"].value = "date"
        self._sheet["I1"].alignment = center_alignment
        self._sheet["I1"].border = border
        self._sheet.merge_cells("I1:I4")
        self._sheet.column_dimensions["I"].width = 15
